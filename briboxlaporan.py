import telebot
import time
import datetime
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from telebot import types

# Setup logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO)
logger = logging.getLogger(__name__)

# Inisialisasi bot dengan token Anda
bot = telebot.TeleBot("8051268487:AAH-lbNqxcQyKj_3rrK3dl4OqwMUKg51jvs")

# Dictionary untuk menyimpan data sementara user
user_data = {}

# List untuk menyimpan semua laporan yang sudah dibuat
all_reports = []


# Time greetings based on current time
def get_greeting():
    now = datetime.datetime.now()
    hour = now.hour

    if 5 <= hour < 12:
        return "Pagi"
    elif 12 <= hour < 15:
        return "Siang"
    elif 15 <= hour < 19:
        return "Sore"
    else:
        return "Malam"


def create_excel_report():
    """Membuat file Excel dari semua laporan yang tersimpan"""
    if not all_reports:
        return None
    
    try:
        # Buat workbook baru
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return None
        ws.title = "Laporan Maintenance"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Definisi header kolom
        headers = [
            "No", "Tanggal", "Jenis Laporan", "Unit Kerja/Divisi", 
            "Kantor Cabang/Direktorat", "Jenis Pekerjaan", "Berangkat", 
            "Tiba", "Mulai", "Selesai", "Serial Number", "Jenis Perangkat", 
            "Type", "Merk", "Progress", "PIC", "No Telepon", "Status", "Waktu Dibuat"
        ]
        
        # Tulis header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if cell is not None:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
    
        # Sort laporan berdasarkan tanggal (dari atas ke bawah / terbaru ke terlama)
        sorted_reports = sorted(all_reports, key=lambda x: datetime.datetime.strptime(x.get('tanggal', '01/01/2000'), '%d/%m/%Y'), reverse=True)
        
        # Tulis data laporan
        for idx, report in enumerate(sorted_reports, 2):
            row_data = [
                idx - 1,  # Nomor urut
                report.get('tanggal', ''),
                report.get('jenis_laporan', ''),
                report.get('unit_kerja', ''),
                report.get('cabang', ''),
                report.get('jenis_pekerjaan', ''),
                report.get('berangkat', ''),
                report.get('tiba', ''),
                report.get('mulai', ''),
                report.get('selesai', ''),
                report.get('serial_number', ''),
                report.get('jenis_perangkat', ''),
                report.get('type', ''),
                report.get('merk', ''),
                report.get('progress', ''),
                report.get('pic', ''),
                report.get('telepon', ''),
                report.get('status', ''),
                report.get('created_at', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                if cell is not None:
                    cell.border = thin_border
                    if col == 1:  # Nomor urut - center alignment
                        cell.alignment = center_alignment
    
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            try:
                column_cells = ws[column_letter]
                if column_cells is not None:
                    for row in column_cells:
                        try:
                            if row and row.value and len(str(row.value)) > max_length:
                                max_length = len(str(row.value))
                        except:
                            pass
            except:
                pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            if hasattr(ws, 'column_dimensions') and ws.column_dimensions is not None:
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Simpan file
        filename = f"Laporan_Maintenance_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    except Exception as e:
        print(f"Error creating Excel report: {e}")
        return None


def save_report_to_list(report_data, jenis_laporan):
    """Menyimpan laporan ke list untuk export Excel"""
    report_entry = report_data.copy()
    report_entry['jenis_laporan'] = jenis_laporan
    report_entry['created_at'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    all_reports.append(report_entry)


def format_summary_cm(data):
    """Format laporan Corrective Maintenance"""
    greeting = get_greeting()

    summary = (
        f"LAPORAN PEKERJAAN CORRECTIVE MAINTENANCE\n\n"
        f"Selamat {greeting} Petugas Call Center, Update Pekerjaan\n\n"
        f"Unit Kerja/Divisi : {data.get('unit_kerja', '')}\n"
        f"Kantor Cabang/Direktorat: {data.get('cabang', '')}\n"
        f"Tanggal : {data.get('tanggal', '')}\n\n"
        f"Jenis Pekerjaan (Problem) : {data.get('jenis_pekerjaan', '')}\n\n"
        f"Berangkat : {data.get('berangkat', '')}\n"
        f"Tiba : {data.get('tiba', '')}\n"
        f"Mulai : {data.get('mulai', '')}\n"
        f"Selesai : {data.get('selesai', '')}\n\n"
        f"Serial Number : {data.get('serial_number', '')}\n"
        f"Jenis perangkat : {data.get('jenis_perangkat', '')}\n"
        f"Type : {data.get('type', '')}\n"
        f"Merk : {data.get('merk', '')}\n\n"
        f"Progress :\n{data.get('progress', '')}\n\n"
        f"PIC : {data.get('pic', '')}\n"
        f"No telepon : {data.get('telepon', '')}\n\n"
        f"Status: {data.get('status', '')}")

    return summary


def format_summary_pm(data):
    """Format laporan Preventive Maintenance"""
    greeting = get_greeting()

    summary = (
        f"LAPORAN PEKERJAAN PREVENTIVE MAINTENANCE\n\n"
        f"Selamat {greeting} Petugas Call Center, Update Pekerjaan\n\n"
        f"Unit Kerja/Divisi : {data.get('unit_kerja', '')}\n"
        f"Kantor Cabang/Direktorat: {data.get('cabang', '')}\n"
        f"Tanggal : {data.get('tanggal', '')}\n\n"
        f"Jenis Pekerjaan (Problem) : {data.get('jenis_pekerjaan', '')}\n\n"
        f"Berangkat : {data.get('berangkat', '')}\n"
        f"Tiba : {data.get('tiba', '')}\n"
        f"Mulai : {data.get('mulai', '')}\n"
        f"Selesai : {data.get('selesai', '')}\n\n"
        f"Progress :\n{data.get('progress', '')}\n\n"
        f"PIC : {data.get('pic', '')}\n"
        f"No telepon : {data.get('telepon', '')}\n\n"
        f"Status: {data.get('status', '')}")

    return summary


def format_summary_tambahan(data):
    """Format laporan Pekerjaan Tambahan"""
    greeting = get_greeting()

    summary = (
        f"LAPORAN PEKERJAAN TAMBAHAN\n\n"
        f"Selamat {greeting} Petugas Call Center, Update Pekerjaan\n\n"
        f"Unit Kerja/Divisi : {data.get('unit_kerja', '')}\n"
        f"Kantor Cabang/Direktorat: {data.get('cabang', '')}\n"
        f"Tanggal : {data.get('tanggal', '')}\n\n"
        f"Jenis Pekerjaan (Problem) : {data.get('jenis_pekerjaan', '')}\n\n"
        f"Berangkat : {data.get('berangkat', '')}\n"
        f"Tiba : {data.get('tiba', '')}\n"
        f"Mulai : {data.get('mulai', '')}\n"
        f"Selesai : {data.get('selesai', '')}\n\n"
        f"Progress :\n{data.get('progress', '')}\n\n"
        f"PIC : {data.get('pic', '')}\n"
        f"No telepon : {data.get('telepon', '')}\n\n"
        f"Status: {data.get('status', '')}")

    return summary


def format_summary_cm_vms(data):
    """Format laporan Corrective Maintenance VMS UKER"""
    greeting = get_greeting()

    summary = (
        f"LAPORAN PEKERJAAN CORRECTIVE MAINTENANCE VMS UKER\n\n"
        f"Selamat {greeting} Petugas Call Center, Update Pekerjaan\n\n"
        f"Unit Kerja/Divisi : {data.get('unit_kerja', '')}\n"
        f"Kantor Cabang/Direktorat: {data.get('cabang', '')}\n"
        f"Tanggal : {data.get('tanggal', '')}\n\n"
        f"Jenis Pekerjaan (Problem) : {data.get('jenis_pekerjaan', '')}\n\n"
        f"Berangkat : {data.get('berangkat', '')}\n"
        f"Tiba : {data.get('tiba', '')}\n"
        f"Mulai : {data.get('mulai', '')}\n"
        f"Selesai : {data.get('selesai', '')}\n\n"
        f"Progress :\n{data.get('progress', '')}\n\n"
        f"PIC : {data.get('pic', '')}\n"
        f"No telepon : {data.get('telepon', '')}\n\n"
        f"Status: {data.get('status', '')}")

    return summary


def format_summary_pm_vms(data):
    """Format laporan Preventive Maintenance VMS UKER"""
    greeting = get_greeting()

    summary = (
        f"LAPORAN PEKERJAAN PREVENTIVE MAINTENANCE VMS UKER\n\n"
        f"Selamat {greeting} Petugas Call Center, Update Pekerjaan\n\n"
        f"Unit Kerja/Divisi : {data.get('unit_kerja', '')}\n"
        f"Kantor Cabang/Direktorat: {data.get('cabang', '')}\n"
        f"Tanggal : {data.get('tanggal', '')}\n\n"
        f"Jenis Pekerjaan (Problem) : {data.get('jenis_pekerjaan', '')}\n\n"
        f"Berangkat : {data.get('berangkat', '')}\n"
        f"Tiba : {data.get('tiba', '')}\n"
        f"Mulai : {data.get('mulai', '')}\n"
        f"Selesai : {data.get('selesai', '')}\n\n"
        f"Progress :\n{data.get('progress', '')}\n\n"
        f"PIC : {data.get('pic', '')}\n"
        f"No telepon : {data.get('telepon', '')}\n\n"
        f"Status: {data.get('status', '')}")

    return summary


@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    """Handler untuk command start dan help"""
    chat_id = message.chat.id

    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True,
                                       resize_keyboard=True,
                                       row_width=2)
    markup.add('🔧 CM', '🛠️ PM', '➕ Tambahan')
    markup.add('📹 CM VMS', '🎥 PM VMS')
    markup.add('❌ Batal', '🚪 Keluar')
    markup.add('ℹ️ Info')
    bot.send_message(
        chat_id,
        "👋 hiii! Silakan pilih jenis laporan yang ingin dibuat:\n\n"
        "• 🔧 CM - Corrective Maintenance\n"
        "• 🛠️ PM - Preventive Maintenance\n"
        "• ➕ Tambahan - Pekerjaan Tambahan\n"
        "• 📹 CM VMS - Corrective Maintenance VMS UKER\n"
        "• 🎥 PM VMS - Preventive Maintenance VMS UKER",
        reply_markup=markup)


@bot.message_handler(func=lambda message: message.text == '🔧 CM')
def start_cm(message):
    """Memulai proses laporan Corrective Maintenance"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'step': 'unit_kerja',
        'jenis_laporan': 'CM',
        'report_type': 'CORRECTIVE MAINTENANCE'
    }

    greeting = get_greeting()
    bot.send_message(chat_id, f"🔧 Selamat {greeting} Petugas Call Center!\n"
                     "Laporan Pekerjaan Corrective Maintenance (CM).\n\n"
                     "📝 Unit Kerja/Divisi:",
                     reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(func=lambda message: message.text == '🛠️ PM')
def start_pm(message):
    """Memulai proses laporan Preventive Maintenance"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'step': 'unit_kerja',
        'jenis_laporan': 'PM',
        'report_type': 'PREVENTIVE MAINTENANCE'
    }

    greeting = get_greeting()
    bot.send_message(chat_id, f"🛠️ Selamat {greeting} Petugas Call Center!\n"
                     "Laporan Pekerjaan Preventive Maintenance (PM).\n\n"
                     "📝 Unit Kerja/Divisi:",
                     reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(func=lambda message: message.text == '➕ Tambahan')
def start_tambahan(message):
    """Memulai proses laporan Pekerjaan Tambahan"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'step': 'unit_kerja',
        'jenis_laporan': 'TAMBAHAN',
        'report_type': 'PEKERJAAN TAMBAHAN'
    }

    greeting = get_greeting()
    bot.send_message(chat_id, f"➕ Selamat {greeting} Petugas Call Center!\n"
                     "Laporan Pekerjaan Tambahan.\n\n"
                     "📝 Unit Kerja/Divisi:",
                     reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(func=lambda message: message.text == '📹 CM VMS')
def start_cm_vms(message):
    """Memulai proses laporan Corrective Maintenance VMS UKER"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'step': 'unit_kerja',
        'jenis_laporan': 'CM_VMS',
        'report_type': 'CORRECTIVE MAINTENANCE VMS UKER'
    }

    greeting = get_greeting()
    bot.send_message(
        chat_id, f"📹 Selamat {greeting} Petugas Call Center!\n"
        "Laporan Pekerjaan Corrective Maintenance VMS UKER.\n\n"
        "📝 Unit Kerja/Divisi:",
        reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(func=lambda message: message.text == '🎥 PM VMS')
def start_pm_vms(message):
    """Memulai proses laporan Preventive Maintenance VMS UKER"""
    chat_id = message.chat.id
    user_data[chat_id] = {
        'step': 'unit_kerja',
        'jenis_laporan': 'PM_VMS',
        'report_type': 'PREVENTIVE MAINTENANCE VMS UKER'
    }

    greeting = get_greeting()
    bot.send_message(
        chat_id, f"🎥 Selamat {greeting} Petugas Call Center!\n"
        "Laporan Pekerjaan Preventive Maintenance VMS UKER.\n\n"
        "📝 Unit Kerja/Divisi:",
        reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(func=lambda message: message.text == '❌ Batal')
def cancel_process(message):
    """Membatalkan proses yang sedang berjalan"""
    chat_id = message.chat.id
    if chat_id in user_data:
        del user_data[chat_id]

    show_main_menu(chat_id, "❌ Proses dibatalkan. Silakan exit bot")


@bot.message_handler(commands=['cancel'])
def cancel_command(message):
    """Command untuk membatalkan proses"""
    chat_id = message.chat.id
    if chat_id in user_data:
        del user_data[chat_id]

    show_main_menu(chat_id, "❌ Proses dibatalkan. Silakan exit bot")


@bot.message_handler(func=lambda message: message.text == '🚪 Keluar')
def exit_bot(message):
    """Handler untuk keluar dari bot"""
    chat_id = message.chat.id
    # Clear user data jika ada
    if chat_id in user_data:
        del user_data[chat_id]
    
    bot.send_message(
        chat_id,
        "👋 Terima kasih telah menggunakan Bot Laporan Bribox\n\n"
        "🔄 Ketik /start untuk memulai lagi\n\n"
        "Sampai jumpa! 😊",
        reply_markup=types.ReplyKeyboardRemove()
    )


@bot.message_handler(commands=['exit', 'quit'])
def exit_command(message):
    """Command untuk keluar dari bot"""
    chat_id = message.chat.id
    # Clear user data jika ada
    if chat_id in user_data:
        del user_data[chat_id]
    
    bot.send_message(
        chat_id,
        "👋 Terima kasih sudah menggunakan Bot Laporan Bribox\n\n"
        "🔄 Ketik /start untuk memulai lagi\n\n"
        "Sampai jumpa! 😊",
        reply_markup=types.ReplyKeyboardRemove()
    )


def show_main_menu(chat_id, message_text):
    """Menampilkan menu utama"""
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True,
                                       resize_keyboard=True,
                                       row_width=2)
    markup.add('🔧 CM', '🛠️ PM', '➕ Tambahan')
    markup.add('📹 CM VMS', '🎥 PM VMS')
    markup.add('❌ Batal', '🚪 Keluar')
    markup.add('ℹ️ Info')

    bot.send_message(chat_id, message_text, reply_markup=markup)


@bot.message_handler(func=lambda message: True)
def handle_message(message):
    """Handler utama untuk semua pesan text"""
    chat_id = message.chat.id

    # Jika user tidak dalam proses laporan, tampilkan menu
    if chat_id not in user_data:
        show_main_menu(chat_id,
                       "📋 Silakan pilih jenis laporan terlebih dahulu:")
        return

    current_step = user_data[chat_id].get('step')
    jenis_laporan = user_data[chat_id].get('jenis_laporan', 'CM')

    # Process berdasarkan step
    if current_step == 'unit_kerja':
        user_data[chat_id]['unit_kerja'] = message.text
        user_data[chat_id]['step'] = 'cabang'
        bot.send_message(chat_id, "🏢 Kantor Cabang/Direktorat:")

    elif current_step == 'cabang':
        user_data[chat_id]['cabang'] = message.text
        user_data[chat_id]['step'] = 'tanggal'
        today = datetime.date.today().strftime("%d/%m/%Y")
        user_data[chat_id]['tanggal'] = today
        bot.send_message(chat_id, f"📅 Tanggal (DD/MM/YYYY) - cth ({today}):")

    elif current_step == 'tanggal':
        if message.text.strip():
            user_data[chat_id]['tanggal'] = message.text
        user_data[chat_id]['step'] = 'jenis_pekerjaan'
        bot.send_message(chat_id, "🔧 Jenis Pekerjaan (Problem):")

    elif current_step == 'jenis_pekerjaan':
        user_data[chat_id]['jenis_pekerjaan'] = message.text
        user_data[chat_id]['step'] = 'berangkat'
        bot.send_message(chat_id, "⏰ Waktu Berangkat (HH:MM):")

    elif current_step == 'berangkat':
        user_data[chat_id]['berangkat'] = message.text
        user_data[chat_id]['step'] = 'tiba'
        bot.send_message(chat_id, "⏰ Waktu Tiba (HH:MM):")

    elif current_step == 'tiba':
        user_data[chat_id]['tiba'] = message.text
        user_data[chat_id]['step'] = 'mulai'
        bot.send_message(chat_id, "⏰ Waktu Mulai (HH:MM):")

    elif current_step == 'mulai':
        user_data[chat_id]['mulai'] = message.text
        user_data[chat_id]['step'] = 'selesai'
        bot.send_message(chat_id, "⏰ Waktu Selesai (HH:MM):")

    elif current_step == 'selesai':
        user_data[chat_id]['selesai'] = message.text
        user_data[chat_id]['step'] = 'serial_number'
        bot.send_message(chat_id, "🔢 Serial Number:")

    elif current_step == 'serial_number':
        user_data[chat_id]['serial_number'] = message.text
        user_data[chat_id]['step'] = 'jenis_perangkat'
        bot.send_message(chat_id, "💻 Jenis Perangkat:")

    elif current_step == 'jenis_perangkat':
        user_data[chat_id]['jenis_perangkat'] = message.text
        user_data[chat_id]['step'] = 'type'
        bot.send_message(chat_id, "📋 Type:")

    elif current_step == 'type':
        user_data[chat_id]['type'] = message.text
        user_data[chat_id]['step'] = 'merk'
        bot.send_message(chat_id, "🏷️ Merk:")

    elif current_step == 'merk':
        user_data[chat_id]['merk'] = message.text
        user_data[chat_id]['step'] = 'progress'
        bot.send_message(chat_id, "📊 Progress:")

    elif current_step == 'progress':
        user_data[chat_id]['progress'] = message.text
        user_data[chat_id]['step'] = 'pic'
        bot.send_message(chat_id, "👤 PIC:")

    elif current_step == 'pic':
        user_data[chat_id]['pic'] = message.text
        user_data[chat_id]['step'] = 'telepon'
        bot.send_message(chat_id, "📞 No Telepon:")

    elif current_step == 'telepon':
        user_data[chat_id]['telepon'] = message.text
        user_data[chat_id]['step'] = 'status'

        # Create keyboard for status
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True,
                                           resize_keyboard=True)
        markup.add('✅ Selesai', '⏳ Pending', '❌ Batal')
        bot.send_message(chat_id, "📊 Status:", reply_markup=markup)

    elif current_step == 'status':
        user_data[chat_id]['status'] = message.text
        user_data[chat_id]['step'] = 'confirmation'

        # Show summary berdasarkan jenis laporan
        if jenis_laporan == 'PM':
            summary = format_summary_pm(user_data[chat_id])
            laporan_type = "Preventive Maintenance"
        elif jenis_laporan == 'TAMBAHAN':
            summary = format_summary_tambahan(user_data[chat_id])
            laporan_type = "Pekerjaan Tambahan"
        elif jenis_laporan == 'CM_VMS':
            summary = format_summary_cm_vms(user_data[chat_id])
            laporan_type = "Corrective Maintenance VMS UKER"
        elif jenis_laporan == 'PM_VMS':
            summary = format_summary_pm_vms(user_data[chat_id])
            laporan_type = "Preventive Maintenance VMS UKER"
        else:
            summary = format_summary_cm(user_data[chat_id])
            laporan_type = "Corrective Maintenance"

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True,
                                           resize_keyboard=True)
        markup.add('✅ Ya', '❌ Tidak')
        bot.send_message(chat_id, f"📋 REVIEW LAPORAN {laporan_type}\n\n"
                         f"{summary}\n\n"
                         "✅ Apakah data sudah benar?",
                         reply_markup=markup)

    elif current_step == 'confirmation':
        if message.text.lower() in ['✅ ya', 'ya', 'yes', 'y']:
            # Format summary berdasarkan jenis laporan
            if jenis_laporan == 'PM':
                summary = format_summary_pm(user_data[chat_id])
                laporan_type = "Preventive Maintenance"
            elif jenis_laporan == 'TAMBAHAN':
                summary = format_summary_tambahan(user_data[chat_id])
                laporan_type = "Pekerjaan Tambahan"
            elif jenis_laporan == 'CM_VMS':
                summary = format_summary_cm_vms(user_data[chat_id])
                laporan_type = "Corrective Maintenance VMS UKER"
            elif jenis_laporan == 'PM_VMS':
                summary = format_summary_pm_vms(user_data[chat_id])
                laporan_type = "Preventive Maintenance VMS UKER"
            else:
                summary = format_summary_cm(user_data[chat_id])
                laporan_type = "Corrective Maintenance"

            bot.send_message(chat_id,
                             f"✅ LAPORAN {laporan_type} BERHASIL DIBUAT!\n\n"
                             f"{summary}",
                             reply_markup=types.ReplyKeyboardRemove())

            # Simpan laporan ke list untuk export Excel
            save_report_to_list(user_data[chat_id], jenis_laporan)

            # Log ke console
            logger.info(f"Laporan {jenis_laporan} dibuat untuk user {chat_id}")

            # Tawarkan export ke Excel
            markup = types.InlineKeyboardMarkup()
            export_btn = types.InlineKeyboardButton("📊 Export ke Excel", callback_data="export_excel")
            markup.add(export_btn)
            
            bot.send_message(
                chat_id,
                "💾 Laporan telah disimpan ke database!\n"
                f"📝 Total laporan tersimpan: {len(all_reports)}\n\n"
                "Ingin mengunduh semua laporan dalam format Excel?",
                reply_markup=markup
            )

            # Jeda waktu 3 detik setelah laporan berhasil dibuat
            time.sleep(3)

        else:
            bot.send_message(chat_id,
                             "❌ Laporan dibatalkan.",
                             reply_markup=types.ReplyKeyboardRemove())

        # Kembali ke menu utama
        show_main_menu(chat_id, "📋 Silakan pilih jenis laporan:")

        # Clear user data
        if chat_id in user_data:
            del user_data[chat_id]


@bot.callback_query_handler(func=lambda call: call.data == "export_excel")
def handle_excel_export(call):
    """Handler untuk callback button export Excel"""
    chat_id = call.message.chat.id
    
    try:
        # Hapus inline keyboard
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
        
        if not all_reports:
            bot.send_message(chat_id, "❌ Tidak ada laporan untuk di-export!")
            return
        
        bot.send_message(chat_id, "⏳ Membuat file Excel... Mohon tunggu...")
        
        # Buat file Excel
        filename = create_excel_report()
        
        if filename:
            # Kirim file Excel ke user
            with open(filename, 'rb') as document:
                bot.send_document(
                    chat_id, 
                    document,
                    caption=f"📊 File Excel Laporan Maintenance\n"
                           f"📅 Dibuat: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
                           f"📝 Total data: {len(all_reports)} laporan\n"
                           f"🔄 Data diurutkan berdasarkan tanggal (terbaru ke terlama)"
                )
            
            # Hapus file setelah dikirim
            os.remove(filename)
            
            bot.send_message(
                chat_id, 
                "✅ File Excel berhasil dibuat dan dikirim!\n\n"
                "📋 Kembali ke menu utama untuk membuat laporan baru."
            )
        else:
            bot.send_message(chat_id, "❌ Gagal membuat file Excel!")
            
    except Exception as e:
        bot.send_message(chat_id, f"❌ Error saat membuat Excel: {str(e)}")
        logger.error(f"Error saat export Excel: {e}")


@bot.message_handler(commands=['export'])
def export_command(message):
    """Command untuk export Excel semua laporan"""
    chat_id = message.chat.id
    
    if not all_reports:
        bot.send_message(
            chat_id, 
            "❌ Belum ada laporan yang tersimpan!\n"
            "📝 Buat laporan terlebih dahulu dengan /start"
        )
        return
    
    markup = types.InlineKeyboardMarkup()
    export_btn = types.InlineKeyboardButton("📊 Export ke Excel", callback_data="export_excel")
    markup.add(export_btn)
    
    bot.send_message(
        chat_id,
        f"📊 **EXPORT LAPORAN KE EXCEL**\n\n"
        f"📝 Total laporan tersimpan: {len(all_reports)}\n"
        f"🔄 Data akan diurutkan berdasarkan tanggal\n\n"
        f"Klik tombol di bawah untuk mengunduh:",
        reply_markup=markup,
        parse_mode='Markdown'
    )


@bot.message_handler(commands=['status'])
def check_status(message):
    """Command untuk mengecek status bot"""
    chat_id = message.chat.id
    bot.send_message(
        chat_id, "🤖 Bot Laporan Maintenance sedang berjalan!\n"
        "✅ Status: Active\n"
        "📊 Fitur: 5 Jenis Laporan\n"
        "• 🔧 Corrective Maintenance\n"
        "• 🛠️ Preventive Maintenance\n"
        "• ➕ Pekerjaan Tambahan\n"
        "• 📹 CM VMS UKER\n"
        "• 🎥 PM VMS UKER\n"
        "ℹ️ Ketik /start untuk memulai")


@bot.message_handler(commands=['info'])
def show_info(message):
    """Command untuk menampilkan info bot"""
    chat_id = message.chat.id
    bot.send_message(
        chat_id, "ℹ️ INFORMASI BOT LAPORAN BRIBOX\n\n"
        "📋 JENIS LAPORAN YANG DIDUKUNG:\n"
        "• 🔧 CM - Corrective Maintenance\n"
        "• 🛠️ PM - Preventive Maintenance\n"
        "• ➕ Tambahan - Pekerjaan Tambahan\n"
        "• 📹 CM VMS - Corrective Maintenance VMS UKER\n"
        "• 🎥 PM VMS - Preventive Maintenance VMS UKER\n\n"
        "⚡ COMMANDS:\n"
        "/start - Memulai bot\n"
        "/cancel - Membatalkan proses\n"
        "/status - Status bot\n"
        "/info - Informasi bot\n"
        "/export - Export laporan ke Excel\n"
        "/exit - Keluar dari bot\n"
        "/quit - Keluar dari bot\n\n"
        "📞 make by verdi")


if __name__ == '__main__':
    print("Bot Laporan Maintenance Multi-Jenis sedang berjalan...")
    print("Supported reports: CM, PM, Tambahan, CM VMS, PM VMS")
    print("Bot ready to receive messages...")
    try:
        bot.polling()
    except Exception as e:
        print(f"❌ Error: {e}")
        print("🔄 Restarting bot...")

